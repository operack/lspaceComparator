import openpyxl


def identify_report_type(excel_file):
    """
    Detect the type of report based on known Full Circle and WMS markers.

    Returns:
        str: One of:
            - "Open Orders from Full Circle"
            - "Shipped Orders from Full Circle"
            - "Open Orders from WMS"
            - "Shipped Orders from WMS"
            - "Unknown Report"
    """
    try:
        workbook = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
        sheet = workbook.active

        a1_value = str(sheet['A1'].value or "").strip()

        # Full Circle reports
        if "Pick Slip Report for Released Picks" in a1_value:
            return "Open Orders from Full Circle"
        if "Daily Shipment Report" in a1_value:
            return "Shipped Orders from Full Circle"

        # WMS reports — detect by header fields
        headers = [str(cell.value).strip() if cell.value else "" for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        if "ShippingInfoId" in headers:
            return "Shipped Orders from WMS"
        if "OrdId" in headers:
            return "Open Orders from WMS"

        return "Unknown Report"

    except Exception as e:
        return f"Error reading file: {str(e)}"


def extract_pick_numbers_from_excel(excel_file):
    """
    Detects the report type and extracts normalized pick numbers
    based on report structure rules.

    Returns:
        dict with:
            - report_type: "OPEN" or "SHIPPED"
            - source: "FULL_CIRCLE" or "WMS"
            - pick_numbers: List[str]
    """
    workbook = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
    sheet = workbook.active

    a1 = str(sheet['A1'].value or "").strip()
    headers = [str(cell.value).strip() if cell.value else "" for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

    # Detect report type and extraction rules
    if "Pick Slip Report for Released Picks" in a1:
        report_type = "OPEN"
        source = "FULL_CIRCLE"
        col_index = 0  # Column A
        start_row = 6
    elif "Daily Shipment Report" in a1:
        report_type = "SHIPPED"
        source = "FULL_CIRCLE"
        col_index = 13  # Column N
        start_row = 6
    elif "ShippingInfoId" in headers:
        report_type = "SHIPPED"
        source = "WMS"
        col_index = 3  # Column D
        start_row = 2
    elif "OrdId" in headers:
        report_type = "OPEN"
        source = "WMS"
        col_index = 2  # Column C
        start_row = 2
    else:
        raise ValueError("Unknown report format")

    # Extract pick numbers
    pick_numbers = []
    for row in sheet.iter_rows(min_row=start_row):
        cell = row[col_index]
        if cell and cell.value:
            raw = str(cell.value).strip().replace("-", "")
            if raw:
                pick_numbers.append(raw)

    return {
        "report_type": report_type,
        "source": source,
        "pick_numbers": pick_numbers
    }
